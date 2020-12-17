import openpyxl
from openpyxl.styles import PatternFill
import os
def changecolor(excel_file):
    wb=openpyxl.load_workbook(excel_file)
    ws=wb['原始记录']
    fille = PatternFill('solid',fgColor='FFFFFF')
    ws['C9'].fill = fille
    ws['B94'].fill = fille
    wb.save(excel_file)
print(os.getcwd())
# changecolor('参数页复制后.xlsx')