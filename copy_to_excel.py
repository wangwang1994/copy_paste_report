import os,sys
import openpyxl
from openpyxl.drawing.image import Image
from zipfile import ZipFile
origin_excel=input('请输入原始excel文件名:')
baogaobianhao=input('请输入报告编号：')
yangpinbianhao=input('请输入样品编号：')
wb_origin = openpyxl.load_workbook(origin_excel)  # Add file name
sheet_canshu = wb_origin["参数"]  # Add Sheet name
# 由于合并的单元格无法在以下的循环中进行使用，因此需要在这里将需要循环部分的
# 单元格进行拆分，在最后保存之前进行合并就可以了
for i in range(4,16):
    sheet_canshu.unmerge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    sheet_canshu.unmerge_cells(start_row=i, start_column=4, end_row=i, end_column=5)
    sheet_canshu.unmerge_cells(start_row=i, start_column=6, end_row=i, end_column=7)
    sheet_canshu.unmerge_cells(start_row=i, start_column=8, end_row=i, end_column=9)
sheet_yuanshijilu=wb_origin['原始记录']
sheet_yuanshijilu.unmerge_cells('C5:D5')

# File to be pasted into
template = openpyxl.load_workbook("轻型汽油车原始记录模板.xlsx")  # Add file name
temp_sheet_canshu = template["参数"]  # Add Sheet name
temp_sheet_yuanshijilu=template['原始记录']



def copyRange(startCol,startRow,endCol,endRow,sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected


# Paste range
# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createData():
    print("Processing...")
    selectedRange_canshu = copyRange(2, 4, 9, 15, sheet_canshu)
    pastingRange_canshu = pasteRange(2, 4, 9, 15, temp_sheet_canshu,selectedRange_canshu)
    selectedRange_yuanshijilu1=copyRange(3,5,4,5,sheet_yuanshijilu)
    pastingRange_yuanshijilu1=pasteRange(3,5,4,5,temp_sheet_yuanshijilu,selectedRange_yuanshijilu1)
    selectedRange_yuanshijilu2=copyRange(4,9,4,9,sheet_yuanshijilu)
    pastingRange_yuanshijilu2=pasteRange(4,10,4,10,temp_sheet_yuanshijilu,selectedRange_yuanshijilu2)
    selectedRange_yuanshijilu3 = copyRange(4, 11, 4, 11, sheet_yuanshijilu)
    pastingRange_yuanshijilu3 = pasteRange(4, 12, 4, 12, temp_sheet_yuanshijilu, selectedRange_yuanshijilu3)

    for i in range(4, 16):
        temp_sheet_canshu.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        temp_sheet_canshu.merge_cells(start_row=i, start_column=4, end_row=i, end_column=5)
        temp_sheet_canshu.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)
        temp_sheet_canshu.merge_cells(start_row=i, start_column=8, end_row=i, end_column=9)
    temp_sheet_yuanshijilu.merge_cells('C5:D5')
    temp_sheet_yuanshijilu['F43']='报告编号：'+baogaobianhao
    temp_sheet_yuanshijilu['B94']='外观检验照片见'+baogaobianhao+'#光盘 文件夹'
    temp_sheet_yuanshijilu['C9']=yangpinbianhao
    template.save("参数页复制后.xlsx")
    print("Range copied and pasted!")

createData()

wb_xiugaihou=openpyxl.load_workbook('参数页复制后.xlsx')
ws_suicheqingdan=wb_xiugaihou['随车清单']

zip = ZipFile(origin_excel)
zip.extractall()
try:
    os.rename('xl/media/image1.jpeg','xl/media/image1.png')
except:
    pass
img1 = Image('xl/media/image1.png')
img1.height=500
img1.width=500
img1.anchor='A3'
ws_suicheqingdan.add_image(img1)
try:
    os.rename('xl/media/image2.jpeg', 'xl/media/image2.png')
except:
    print('没有第二张清单')
try:
    img2=Image('xl/media/image2.png')
    img2.height = 500
    img2.width = 500
    img2.anchor = 'A18'
    ws_suicheqingdan.add_image(img2)
except FileNotFoundError:
    print('没有第二张清单')
wb_xiugaihou.save(baogaobianhao+'.xlsx')
os.remove('xl/media/image1.png')
try:
    os.remove('xl/media/image2.png')
except:
    pass